"""
Assets sheet bulk plan: Source Path | Target Path only.

Source Path = local page folder containing Desktop / Mobile / Tablet
Target Path = DAM page folder e.g. /content/dam/we-retail/en/men

Preview plans uploads; apply can call DamService.upload_from_local later.
"""
from __future__ import annotations

import io
import re
from typing import List

from openpyxl import load_workbook

from backend.app.services.dam_service import DamService, BREAKPOINTS, normalize_dam_path


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def parse_assets_sheet(content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    rows = []
    errors = []
    sheet_used = None

    for sheet_name in wb.sheetnames:
        if _norm(sheet_name) not in ("assets", "asset", "dam"):
            continue
        ws = wb[sheet_name]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        headers = [str(h).strip() if h is not None else "" for h in data[0]]
        src_i = tgt_i = None
        for i, h in enumerate(headers):
            hl = _norm(h)
            if hl in ("source path", "source", "local path", "local folder"):
                src_i = i
            elif hl in ("target path", "target", "dam path", "dam folder"):
                tgt_i = i
        if src_i is None or tgt_i is None:
            errors.append(f"{sheet_name}: need columns Source Path and Target Path")
            continue
        sheet_used = sheet_name
        for r_i, row in enumerate(data[1:], start=2):
            if not row:
                continue
            src = row[src_i] if src_i < len(row) else None
            tgt = row[tgt_i] if tgt_i < len(row) else None
            if src is None or str(src).strip() == "":
                continue
            if tgt is None or str(tgt).strip() == "":
                errors.append(f"{sheet_name} row {r_i}: Target Path required")
                continue
            # skip italic example-looking placeholder? still process
            rows.append({
                "excel_row": r_i,
                "source_path": str(src).strip(),
                "target_path": str(tgt).strip(),
            })
        break

    return {"status": "success", "sheet": sheet_used, "rows": rows, "errors": errors}


def plan_asset_uploads(content: bytes) -> dict:
    parsed = parse_assets_sheet(content)
    if not parsed.get("rows"):
        return {
            "status": "error",
            "message": "No Assets rows found. Sheet must be named Assets with Source Path and Target Path.",
            "parse": parsed,
        }

    dam = DamService()
    plans = []
    for row in parsed["rows"]:
        entry = {
            "excel_row": row["excel_row"],
            "source_path": row["source_path"],
            "target_path": row["target_path"],
            "errors": [],
            "warnings": [],
            "scan": None,
            "folders": None,
            "upload_plan": [],
        }
        try:
            target = normalize_dam_path(row["target_path"])
            entry["target_path"] = target
        except Exception as e:
            entry["errors"].append(str(e))
            plans.append(entry)
            continue

        entry["folders"] = dam.inspect_page_dam_path(target)
        if not entry["folders"].get("all_ready"):
            entry["warnings"].append(
                "DAM folders missing — will be created on Apply if confirm_create_folders=true"
            )

        scan = dam.scan_local_page_folder(row["source_path"])
        entry["scan"] = {
            "status": scan.get("status"),
            "message": scan.get("message"),
            "total_files": scan.get("total_files"),
            "oversized_count": scan.get("oversized_count"),
            "unmatched_dirs": scan.get("unmatched_dirs"),
        }
        if scan.get("status") != "success":
            entry["errors"].append(scan.get("message") or "Local path missing or has no assets")
            entry["local_exists"] = scan.get("exists")
            entry["scan"] = {
                "status": scan.get("status"),
                "message": scan.get("message"),
                "mode": scan.get("mode"),
            }
            plans.append(entry)
            continue

        entry["mode"] = scan.get("mode")
        if scan.get("mode") == "empty" or not scan.get("total_files"):
            entry["errors"].append(
                "Local folder has no uploadable assets (need Desktop/Mobile/Tablet or files in folder)"
            )
            plans.append(entry)
            continue

        if scan.get("mode") == "flat":
            for item in scan.get("flat_files") or []:
                rec = {
                    "breakpoint": None,
                    "local_path": item.get("local_path"),
                    "original_name": item.get("original_name"),
                    "dam_name": item.get("dam_name"),
                    "dam_path": f"{target}/{item.get('dam_name')}",
                    "size_kb": item.get("size_kb"),
                    "size_allowed": item.get("size_allowed"),
                    "mime": item.get("content_type"),
                    "action": "upload" if item.get("size_allowed") else "reject_size",
                    "message": item.get("size_message") or "Flat mode → DAM page folder",
                }
                entry["upload_plan"].append(rec)
        else:
            for bp in BREAKPOINTS:
                for item in scan.get("breakpoints", {}).get(bp) or []:
                    rec = {
                        "breakpoint": bp,
                        "local_path": item.get("local_path"),
                        "original_name": item.get("original_name"),
                        "dam_name": item.get("dam_name"),
                        "dam_path": f"{target}/{bp}/{item.get('dam_name')}",
                        "size_kb": item.get("size_kb"),
                        "size_allowed": item.get("size_allowed"),
                        "mime": item.get("content_type"),
                        "action": "upload" if item.get("size_allowed") else "reject_size",
                        "message": item.get("size_message"),
                    }
                    entry["upload_plan"].append(rec)

        entry["summary"] = {
            "files": len(entry["upload_plan"]),
            "will_upload": sum(1 for u in entry["upload_plan"] if u["action"] == "upload"),
            "rejected_size": sum(1 for u in entry["upload_plan"] if u["action"] == "reject_size"),
        }
        entry["ootb_note"] = (
            "OOTB single-asset components (e.g. Hero): use desktop DAM path as fileReference. "
            "Mobile/tablet files are still uploaded for custom components / future use."
        )
        plans.append(entry)

    return {
        "status": "success",
        "message": "Assets plan (preview only)",
        "sheet": parsed.get("sheet"),
        "parse_errors": parsed.get("errors") or [],
        "plans": plans,
        "summary": {
            "rows": len(plans),
            "total_planned_uploads": sum(p.get("summary", {}).get("will_upload", 0) for p in plans),
            "total_rejected_size": sum(p.get("summary", {}).get("rejected_size", 0) for p in plans),
        },
    }
