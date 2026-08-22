"""
Pages sheet bulk: Page Path | Create (Y/N) | Template Name

Preview: inspect path, resolve template name → path, report plan
Apply: create missing pages (parents must exist or be earlier rows)
"""
from __future__ import annotations

import io
import re
from typing import Dict, List, Optional

from openpyxl import load_workbook

from backend.app.services.page_service import PageService, normalize_content_path, adobe_page_name


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def parse_pages_sheet(content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    rows = []
    errors = []
    sheet_used = None

    for sheet_name in wb.sheetnames:
        if _norm(sheet_name) not in ("pages", "page", "page creation"):
            continue
        ws = wb[sheet_name]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        headers = [str(h).strip() if h is not None else "" for h in data[0]]
        col = {}
        for i, h in enumerate(headers):
            hl = _norm(h)
            if hl in ("page path", "pagepath", "path"):
                col["path"] = i
            elif hl in ("create (y/n)", "create", "create?", "create yn"):
                col["create"] = i
            elif hl in ("template name", "template", "template path"):
                col["template"] = i
            elif hl in ("title", "page title"):
                col["title"] = i
        if "path" not in col:
            errors.append(f"{sheet_name}: Page Path column required")
            continue
        sheet_used = sheet_name
        for r_i, row in enumerate(data[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            path = row[col["path"]] if col["path"] < len(row) else None
            if path is None or str(path).strip() == "":
                continue
            create_raw = ""
            if "create" in col and col["create"] < len(row) and row[col["create"]] is not None:
                create_raw = str(row[col["create"]]).strip().lower()
            create = create_raw in ("y", "yes", "true", "1")
            template = None
            if "template" in col and col["template"] < len(row) and row[col["template"]]:
                template = str(row[col["template"]]).strip()
            title = None
            if "title" in col and col["title"] < len(row) and row[col["title"]]:
                title = str(row[col["title"]]).strip()
            rows.append({
                "excel_row": r_i,
                "page_path": str(path).strip(),
                "create": create,
                "template_name": template,
                "title": title,
            })
        break

    return {"status": "success", "sheet": sheet_used, "rows": rows, "errors": errors}


def _resolve_template(ps: PageService, parent_path: str, template_name: str) -> Optional[str]:
    """Match CA template name to path (title, name, or full path)."""
    if not template_name:
        return None
    t = template_name.strip()
    if t.startswith("/conf/") or t.startswith("/apps/") or t.startswith("/libs/"):
        return t
    listing = ps.list_allowed_templates(parent_path, walk_ancestors=True)
    templates = listing.get("templates") or []
    tn = _norm(t)
    for item in templates:
        if _norm(item.get("path") or "") == tn:
            return item["path"]
        if _norm(item.get("title") or "") == tn:
            return item["path"]
        if _norm(item.get("name") or "") == tn:
            return item["path"]
        # partial title match
        if tn in _norm(item.get("title") or "") or tn in _norm(item.get("path") or ""):
            return item["path"]
    return None



def preview_pages(content: bytes) -> dict:
    """
    Preview page creates. Parents listed in the same sheet with Create=Y
    are treated as available for child rows (same bulk batch).
    Rows are ordered by path depth so parents plan before children.
    """
    parsed = parse_pages_sheet(content)
    if not parsed.get("rows"):
        return {
            "status": "skipped",
            "message": "No Pages sheet/rows in Excel — skipped (existing pages workflow; pages must already exist).",
            "parse": parsed,
            "plans": [],
            "results": [],
        }

    ps = PageService()
    # Sort shallow → deep so parents appear first
    rows = sorted(
        parsed["rows"],
        key=lambda r: (str(r.get("page_path") or "").count("/"), str(r.get("page_path") or "")),
    )

    # Paths we will create in this batch
    will_create_paths = set()
    for row in rows:
        if row.get("create"):
            try:
                will_create_paths.add(normalize_content_path(row["page_path"]))
            except Exception:
                pass

    plans = []
    for row in rows:
        entry = {
            "excel_row": row["excel_row"],
            "page_path": row["page_path"],
            "create": row["create"],
            "template_name": row.get("template_name"),
            "title": row.get("title"),
            "errors": [],
            "warnings": [],
            "action": "skip",
        }
        try:
            path = normalize_content_path(row["page_path"])
            entry["page_path"] = path
        except Exception as e:
            entry["errors"].append(str(e))
            entry["action"] = "blocked"
            plans.append(entry)
            continue

        insp = ps.inspect_page_path(path)
        entry["inspection"] = {
            "target_exists": insp.get("target_exists"),
            "missing": [m.get("path") for m in (insp.get("missing") or [])],
            "all_ready": insp.get("all_ready"),
        }

        if insp.get("target_exists"):
            entry["action"] = "exists"
            entry["warnings"].append("Page already exists — will not recreate")
            plans.append(entry)
            continue

        if not row["create"]:
            entry["action"] = "skip"
            entry["warnings"].append("Create=N — skipped")
            plans.append(entry)
            continue

        # Missing parents not in this batch → blocked
        missing_parents = [
            m for m in (insp.get("missing") or [])
            if m.get("path") != path
        ]
        unresolved = []
        for m in missing_parents:
            mp = m.get("path")
            if mp in will_create_paths:
                entry["warnings"].append(
                    f"Parent {mp} will be created earlier in this same Excel batch"
                )
            else:
                unresolved.append(mp)

        if unresolved:
            entry["errors"].append(
                "Missing parent segment(s) not in this Excel Create=Y list: "
                + ", ".join(unresolved)
                + " — add parent rows with Create=Y above, or create them first"
            )
            entry["action"] = "blocked"
            plans.append(entry)
            continue

        if not row.get("template_name"):
            entry["errors"].append("Template Name required when Create=Y")
            entry["action"] = "blocked"
            plans.append(entry)
            continue

        parent = path.rsplit("/", 1)[0]
        # For template resolution, use nearest existing ancestor or parent that will exist
        resolve_parent = parent
        if not ps.path_exists(parent):
            # walk up to existing
            cursor = parent
            while cursor.startswith("/content") and not ps.path_exists(cursor):
                if cursor in will_create_paths:
                    break
                nxt = cursor.rsplit("/", 1)[0]
                if nxt == cursor:
                    break
                cursor = nxt
            resolve_parent = cursor if ps.path_exists(cursor) else parent

        resolved = _resolve_template(ps, resolve_parent, row["template_name"])
        if not resolved:
            entry["errors"].append(
                f"Could not resolve template '{row['template_name']}'. "
                "Use a name from Instructions or full /conf/... path."
            )
            entry["action"] = "blocked"
            listing = ps.list_allowed_templates(resolve_parent, walk_ancestors=True)
            entry["available_templates"] = [
                {"title": t.get("title"), "path": t.get("path")}
                for t in (listing.get("templates") or [])[:20]
            ]
            plans.append(entry)
            continue

        entry["template_path"] = resolved
        entry["action"] = "create_page"
        entry["title"] = row.get("title") or path.rsplit("/", 1)[-1].replace("-", " ").title()
        plans.append(entry)

    return {
        "status": "success",
        "message": "Pages plan (preview) — parents in same batch allowed for children",
        "sheet": parsed.get("sheet"),
        "parse_errors": parsed.get("errors") or [],
        "plans": plans,
        "summary": {
            "rows": len(plans),
            "will_create": sum(1 for p in plans if p.get("action") == "create_page"),
            "exists": sum(1 for p in plans if p.get("action") == "exists"),
            "skipped": sum(1 for p in plans if p.get("action") == "skip"),
            "blocked": sum(1 for p in plans if p.get("action") == "blocked"),
        },
    }


def apply_pages(content: bytes, performed_by: str = "system") -> dict:
    """Create pages shallow → deep so parents exist before children."""
    preview = preview_pages(content)
    if preview.get("status") != "success":
        return preview

    # Order create_page by depth
    plans = list(preview.get("plans") or [])
    creates = [p for p in plans if p.get("action") == "create_page"]
    creates.sort(key=lambda p: (str(p.get("page_path") or "").count("/"), str(p.get("page_path") or "")))
    others = [p for p in plans if p.get("action") != "create_page"]

    ps = PageService()
    results = []

    for plan in others:
        results.append({
            "excel_row": plan.get("excel_row"),
            "page_path": plan.get("page_path"),
            "action": plan.get("action"),
            "errors": list(plan.get("errors") or []),
            "status": "skipped" if plan.get("action") != "blocked" else "error",
            "message": plan.get("action"),
        })

    for plan in creates:
        out = {
            "excel_row": plan.get("excel_row"),
            "page_path": plan.get("page_path"),
            "action": "create_page",
            "errors": [],
            "status": "error",
        }
        path = plan["page_path"]
        parent = path.rsplit("/", 1)[0]
        name = path.rsplit("/", 1)[-1]

        # Parent must exist now (created earlier in this loop or already in AEM)
        if not ps.path_exists(parent):
            out["errors"].append(f"Parent still missing at apply time: {parent}")
            out["message"] = "parent_missing"
            results.append(out)
            continue

        r = ps.create_page(
            parent_path=parent,
            name=name,
            title=plan.get("title") or name,
            template_path=plan["template_path"],
            performed_by=performed_by,
        )
        out["result"] = r
        out["status"] = r.get("status")
        out["message"] = r.get("message") or r.get("kind")
        if r.get("status") != "success":
            out["errors"].append(r.get("message") or "Create failed")
        results.append(out)

    ok = sum(1 for r in results if r.get("status") == "success")
    err = sum(1 for r in results if r.get("status") == "error" or r.get("errors"))
    return {
        "status": "success" if err == 0 else "partial",
        "message": f"Pages apply — ok: {ok}, errors: {err}",
        "results": results,
        "performed_by": performed_by,
    }
