from openpyxl import load_workbook
from typing import Dict, List, Any
import io
from backend.app.services.field_mapping import get_real_field_name

class ExcelProcessor:
    """
    Reads the multi-tab Excel template and converts it into
    a clean list of planned updates (for Preview + Apply).
    """

    def process(self, file_bytes: bytes) -> Dict[str, Any]:
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return {
                "status": "error",
                "message": f"Could not read Excel file: {str(e)}"
            }

        result = {
            "status": "success",
            "seo_updates": [],
            "component_updates": [],
            "summary": {
                "total_seo_rows": 0,
                "total_component_rows": 0,
                "sheets_processed": []
            }
        }

        for sheet_name in wb.sheetnames:
            # Skip instruction / help sheets
            lower_name = sheet_name.strip().lower()
            if lower_name in ["instructions", "howtoaddnewcomponent", "readme", "help"]:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows or len(rows) < 2:
                continue

            # First row = headers
            headers = []
            for h in rows[0]:
                if h is None:
                    headers.append(None)
                else:
                    headers.append(str(h).strip())

            # Clean headers (remove None)
            valid_indexes = [i for i, h in enumerate(headers) if h]

            if lower_name == "seo":
                updates = self._process_seo_sheet(rows[1:], headers, valid_indexes)
                result["seo_updates"].extend(updates)
                result["summary"]["total_seo_rows"] += len(updates)
                result["summary"]["sheets_processed"].append(sheet_name)
            else:
                # Treat as a Component sheet
                updates = self._process_component_sheet(sheet_name, rows[1:], headers, valid_indexes)
                result["component_updates"].extend(updates)
                result["summary"]["total_component_rows"] += len(updates)
                result["summary"]["sheets_processed"].append(sheet_name)

        return result

    def _process_seo_sheet(self, data_rows, headers, valid_indexes) -> List[Dict]:
        updates = []
        # Find Page Path column (case-insensitive)
        page_path_idx = None
        for i, h in enumerate(headers):
            if h and h.lower().replace(" ", "") in ["pagepath", "page_path", "path"]:
                page_path_idx = i
                break

        if page_path_idx is None:
            return updates

        for row in data_rows:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            page_path = row[page_path_idx]
            if not page_path or str(page_path).strip() == "":
                continue

            page_path = str(page_path).strip()
            properties = {}

            for i in valid_indexes:
                if i == page_path_idx:
                    continue
                header = headers[i]
                value = row[i] if i < len(row) else None

                if value is None or str(value).strip() == "":
                    continue

                # Use the central mapping
                real_field = get_real_field_name(str(header).strip())
                properties[real_field] = str(value).strip()

            if properties:
                updates.append({
                    "page_path": page_path,
                    "properties": properties
                })

        return updates

    def _process_component_sheet(self, sheet_name, data_rows, headers, valid_indexes) -> List[Dict]:
        updates = []

        # Find important columns
        page_path_idx = None
        instance_idx = None

        for i, h in enumerate(headers):
            if not h:
                continue
            h_clean = h.lower().replace(" ", "").replace("_", "")
            if h_clean in ["pagepath", "page_path", "path"]:
                page_path_idx = i
            elif h_clean in ["instance", "instancenumber", "instance_number", "no", "number"]:
                instance_idx = i

        if page_path_idx is None:
            return updates

        for row in data_rows:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            page_path = row[page_path_idx]
            if not page_path or str(page_path).strip() == "":
                continue

            page_path = str(page_path).strip()

            # Instance number (default 1)
            instance = 1
            if instance_idx is not None and instance_idx < len(row) and row[instance_idx] is not None:
                try:
                    instance = int(row[instance_idx])
                except:
                    instance = 1

            properties = {}
            for i in valid_indexes:
                if i == page_path_idx or i == instance_idx:
                    continue
                header = headers[i]
                value = row[i] if i < len(row) else None

                if value is None or str(value).strip() == "":
                    continue

                # Convert friendly Excel name → real AEM field name
                real_field = get_real_field_name(str(header).strip())
                properties[real_field] = str(value).strip()

            if properties:
                updates.append({
                    "page_path": page_path,
                    "component_name": sheet_name.strip(),   # tab name = component name
                    "instance": instance,
                    "properties": properties
                })

        return updates