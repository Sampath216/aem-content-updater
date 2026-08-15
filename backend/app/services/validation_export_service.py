"""Export validation report to Excel (and use JSON as-is from API)."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def validation_report_to_xlsx(report: Dict[str, Any]) -> bytes:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    pass_fill = PatternFill("solid", fgColor="DCFCE7")
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    warn_fill = PatternFill("solid", fgColor="FFEDD5")

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "AEM Bulk Validation Report"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    ws["A3"] = report.get("message") or ""
    ws["A4"] = f"Status: {report.get('status')}"
    summary = report.get("summary") or {}
    ws["A6"] = "Pass"
    ws["B6"] = summary.get("pass", 0)
    ws["A7"] = "Fail"
    ws["B7"] = summary.get("fail", 0)
    ws["A8"] = "Warn"
    ws["B8"] = summary.get("warn", 0)

    hl = report.get("high_level") or {}
    ws["A10"] = "Section"
    ws["B10"] = "Pass"
    ws["C10"] = "Fail"
    ws["D10"] = "Warn"
    for col in range(1, 5):
        ws.cell(10, col).fill = header_fill
        ws.cell(10, col).font = header_font
    r = 11
    for name in ("assets", "pages", "components", "seo"):
        block = hl.get(name) or {}
        ws.cell(r, 1, name)
        ws.cell(r, 2, block.get("pass", 0))
        ws.cell(r, 3, block.get("fail", 0))
        ws.cell(r, 4, block.get("warn", 0))
        r += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12

    detailed = report.get("detailed") or {}
    for section, items in detailed.items():
        name = str(section)[:31]
        wss = wb.create_sheet(name)
        headers = [
            "Result", "Message", "Page Path", "DAM Path", "Component Path",
            "Component", "Size KB", "Excel Row", "Sheet", "Field Checks",
        ]
        for c, h in enumerate(headers, 1):
            cell = wss.cell(1, c, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
        for i, it in enumerate(items or [], start=2):
            sev = it.get("severity") or ""
            fill = pass_fill if sev == "pass" else (warn_fill if sev == "warn" else fail_fill)
            vals = [
                sev,
                it.get("message"),
                it.get("page_path"),
                it.get("dam_path"),
                it.get("component_path"),
                it.get("component"),
                it.get("size_kb"),
                it.get("excel_row"),
                it.get("sheet"),
                "",
            ]
            fcs = it.get("field_checks") or []
            if fcs:
                parts = []
                for fc in fcs:
                    parts.append(
                        f"{fc.get('field')}: expected={fc.get('expected')!s} actual={fc.get('actual')!s} ok={fc.get('ok')}"
                    )
                vals[9] = " | ".join(parts)
            for c, v in enumerate(vals, 1):
                cell = wss.cell(i, c, v if v is not None else "")
                cell.border = thin
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, 11):
            wss.column_dimensions[chr(64 + c) if c <= 26 else "A"].width = 18
        wss.column_dimensions["B"].width = 50
        wss.column_dimensions["J"].width = 60

    # Field-level sheet (exploded)
    wsf = wb.create_sheet("Field Level")
    fh = ["Section", "Page/Component Path", "Field", "Expected", "Actual", "OK", "Message"]
    for c, h in enumerate(fh, 1):
        cell = wsf.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
    row_i = 2
    for section, items in detailed.items():
        for it in items or []:
            fcs = it.get("field_checks") or []
            if not fcs:
                continue
            for fc in fcs:
                ok = bool(fc.get("ok"))
                fill = pass_fill if ok else fail_fill
                vals = [
                    section,
                    it.get("component_path") or it.get("page_path") or it.get("dam_path"),
                    fc.get("field") or fc.get("jcr_field"),
                    fc.get("expected"),
                    fc.get("actual"),
                    "Y" if ok else "N",
                    it.get("message"),
                ]
                for c, v in enumerate(vals, 1):
                    cell = wsf.cell(row_i, c, v if v is not None else "")
                    cell.fill = fill
                    cell.border = thin
                row_i += 1
    for c, w in enumerate([12, 45, 18, 30, 30, 8, 40], 1):
        wsf.column_dimensions[chr(64 + c)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
