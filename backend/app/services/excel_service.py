from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from backend.app.services.aem_client import AEMClient

class ExcelTemplateService:

    def generate_template(self, components: list) -> BytesIO:
        """
        components = list of dicts with:
        {
            "resourceType": "weretail/components/content/heroimage",
            "label": "Hero Image"   (optional)
        }
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Content Template"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Fixed columns that every template will have
        base_columns = ["Page Path", "Component Name", "Component Path"]
        
        # Collect all unique fields from the selected components
        all_fields = []
        component_fields_map = {}

        aem = AEMClient()

        for comp in components:
            resource_type = comp.get("resourceType")
            label = comp.get("label") or resource_type.split("/")[-1]

            # We need a sample component path to discover fields.
            # For template generation we will use a temporary approach:
            # We ask the caller to also send one real component path of that type
            # OR we just use common field discovery via dialog.
            
            # For now we use dialog discovery only (safer for template)
            fields = aem._get_dialog_fields(resource_type)
            
            # Clean fields
            clean_fields = []
            for f in fields:
                f = f.lstrip("./")
                if f and f not in clean_fields:
                    clean_fields.append(f)

            component_fields_map[label] = clean_fields
            for f in clean_fields:
                if f not in all_fields:
                    all_fields.append(f)

        # Final column order
        columns = base_columns + all_fields

        # Write header
        for col_idx, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Add a few empty example rows so client understands the format
        for row in range(2, 6):
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row, column=col_idx, value="")
                cell.border = thin_border

        # Add instruction sheet
        ws2 = wb.create_sheet("Instructions")
        ws2["A1"] = "How to use this template"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2["A3"] = "1. Fill the 'Page Path' column (example: /content/we-retail/ca/en/men)"
        ws2["A4"] = "2. Fill the 'Component Name' column with the name of the component (example: heroimage)"
        ws2["A5"] = "3. Leave 'Component Path' empty if you want the system to find it, or put the exact path"
        ws2["A6"] = "4. Fill the remaining columns with the content for each field"
        ws2["A8"] = "Important:"
        ws2["A9"] = "- One row = one component instance"
        ws2["A10"] = "- You can have multiple rows for the same component type"
        ws2["A11"] = "- Boolean fields should be true or false"
        ws2["A12"] = "- Do not change the header row"

        # Adjust column widths
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        ws2.column_dimensions["A"].width = 80

        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output