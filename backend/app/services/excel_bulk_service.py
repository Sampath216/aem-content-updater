"""
Excel bulk pipeline: parse → dedupe → validate all → apply.

Page Properties rows update jcr:content (not a component search by name).
Component rows resolve instance by resourceType match + 1-based instance index.
"""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from backend.app.services.dictionary_service import load_dictionary, resolve_label
from backend.app.services.aem_client import AEMClient

def _resolve_resource_type_from_name(name: str, dictionary: dict) -> Optional[str]:
    """Map sheet/component label (Title, button, Hero Image) → sling resourceType."""
    n = _norm(name or "")
    if not n:
        return None
    # exact resourceType string
    if "/" in (name or "") and dictionary and name in dictionary:
        return name
    best = None
    for rt, meta in (dictionary or {}).items():
        if not isinstance(meta, dict):
            continue
        label = _norm(meta.get("label") or "")
        leaf = _norm(rt.split("/")[-1] if rt else "")
        if n == label or n == leaf or n.replace(" ", "") == leaf.replace(" ", ""):
            return rt
        if leaf in n or n in leaf:
            best = best or rt
    return best


def _resolve_field_name(header: str, resource_type: Optional[str], dictionary: dict) -> str:
    """CA header → dialog field name."""
    h = (header or "").strip()
    hl = _norm(h)
    if not h:
        return h
    # Common title/heading mappings (dynamic-friendly defaults)
    if hl in ("title", "heading title", "component title") and resource_type and "title" in _norm(resource_type):
        return "jcr:title"
    if hl in ("title",) and resource_type and "title" in _norm(resource_type or ""):
        return "jcr:title"
    if hl in ("type", "type / size", "type/size", "heading level", "heading type", "size"):
        if resource_type and "title" in _norm(resource_type):
            return "type"
    resolved = None
    if resource_type:
        try:
            resolved = resolve_label(resource_type, h)
        except Exception:
            resolved = None
    if not resolved and resource_type and dictionary:
        fields = (dictionary.get(resource_type) or {}).get("fields") or {}
        for fn, aliases in fields.items():
            alist = aliases if isinstance(aliases, list) else [str(aliases)]
            if _norm(fn) == hl or any(_norm(a) == hl for a in alist):
                resolved = fn
                break
    if not resolved and dictionary:
        for rt, meta in dictionary.items():
            if not isinstance(meta, dict):
                continue
            fields = meta.get("fields") or {}
            for fn, aliases in fields.items():
                alist = aliases if isinstance(aliases, list) else [str(aliases)]
                if _norm(fn) == hl or any(_norm(a) == hl for a in alist):
                    resolved = fn
                    break
            if resolved:
                break
    if not resolved:
        # Title sheet column "Title" → jcr:title when unresolved
        if hl == "title":
            return "jcr:title"
        resolved = h
    return resolved


def normalize_write_value(field_name: str, value):
    if value is None:
        return value
    s = str(value).strip()
    fn = (field_name or "").lower().replace(" ", "")
    truthy = {"y", "yes", "true", "1", "on"}
    falsy = {"n", "no", "false", "0", "off"}
    sl = s.lower()
    if sl in truthy | falsy:
        if any(h in fn for h in ("fullwidth", "checkbox", "boolean", "enabled", "hide", "openinnew")) or fn in ("usefullwidth",):
            return "true" if sl in truthy else "false"
    if fn in ("type", "titletype", "headingtype", "size") or "typesize" in fn:
        if len(s) >= 2 and s[0].upper() == "H" and s[1:].isdigit():
            return s.lower()
    return value


def normalize_props(props: dict) -> dict:
    if not props:
        return {}
    return {k: normalize_write_value(str(k), v) for k, v in props.items()}



def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _is_meta_header(h: str) -> bool:
    return str(h or "").strip().startswith("__resourceType__")


def _parse_resource_type_from_sheet(ws) -> Optional[str]:
    """Read hidden __resourceType__=... from header row."""
    for cell in ws[1]:
        val = str(cell.value or "")
        if val.startswith("__resourceType__="):
            return val.split("=", 1)[1].strip()
    return None


def _sheet_is_page_properties(sheet_name: str, resource_type: Optional[str]) -> bool:
    sn = _norm(sheet_name)
    rt = _norm(resource_type or "")
    if rt in ("page_properties", "cq:page/jcr:content"):
        return True
    if "page properties" in sn or sn in ("seo", "page properties / seo", "page properties  seo"):
        return True
    if rt.endswith("/structure/page"):
        return True
    return False


def _row_key_page(page_path: str) -> str:
    return _norm(page_path.rstrip("/"))


def _row_key_component(page_path: str, resource_type: str, instance: int) -> str:
    return f"{_row_key_page(page_path)}|{_norm(resource_type)}|{instance}"


def parse_workbook(content: bytes) -> dict:
    """
    Parse all sheets into structured rows.
    Returns { status, page_rows, component_rows, duplicates_removed, errors }
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    dictionary = load_dictionary()
    page_rows: List[dict] = []
    component_rows: List[dict] = []
    parse_errors: List[str] = []
    duplicates_removed = 0

    seen_page = {}
    seen_comp = {}

    for sheet_name in wb.sheetnames:
        sn = _norm(sheet_name)
        # Skip non-update sheets (handled by other bulk steps)
        if sn in (
            "instructions",
            "howto",
            "how to use",
            "readme",
            "assets",
            "asset",
            "pages",
            "page",
            "page creation",
        ):
            continue
        if sn.startswith("add ") or sheet_name.strip().lower().startswith("add "):
            continue  # component ADD — separate pipeline
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        resource_type = _parse_resource_type_from_sheet(ws)
        if not resource_type:
            resource_type = _resolve_resource_type_from_name(sheet_name, dictionary)

        # Map header index → dialog field name via dictionary
        header_map = {}  # col_idx -> field_name or special
        for idx, h in enumerate(headers):
            if not h or _is_meta_header(h):
                continue
            hl = _norm(h)
            if hl in ("page path", "pagepath", "path"):
                header_map[idx] = "__page_path__"
            elif hl in ("instance", "instance number", "instance_no"):
                header_map[idx] = "__instance__"
            else:
                if _sheet_is_page_properties(sheet_name, resource_type):
                    resolved = _resolve_field_name(h, "page_properties", dictionary)
                else:
                    resolved = _resolve_field_name(h, resource_type, dictionary)
                header_map[idx] = resolved

        is_page = _sheet_is_page_properties(sheet_name, resource_type)

        for r_i, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            page_path = None
            instance = 1
            props = {}
            for idx, field_key in header_map.items():
                if idx >= len(row):
                    continue
                val = row[idx]
                if val is None or str(val).strip() == "":
                    continue
                if field_key == "__page_path__":
                    page_path = str(val).strip()
                elif field_key == "__instance__":
                    try:
                        instance = int(float(str(val).strip()))
                    except Exception:
                        instance = 1
                else:
                    props[field_key] = val

            if not page_path:
                parse_errors.append(f"{sheet_name} row {r_i}: missing Page Path")
                continue
            if not props:
                continue

            if is_page:
                key = _row_key_page(page_path)
                if key in seen_page:
                    duplicates_removed += 1
                    # later row wins
                    page_rows[:] = [p for p in page_rows if _row_key_page(p["page_path"]) != key]
                seen_page[key] = True
                page_rows.append({
                    "sheet": sheet_name,
                    "excel_row": r_i,
                    "page_path": page_path.rstrip("/"),
                    "properties": props,
                    "resourceType": "page_properties",
                })
            else:
                rt = resource_type or _resolve_resource_type_from_name(sheet_name, dictionary) or sheet_name
                key = _row_key_component(page_path, rt, instance)
                if key in seen_comp:
                    duplicates_removed += 1
                    component_rows[:] = [
                        c for c in component_rows
                        if _row_key_component(c["page_path"], c["resourceType"], c["instance"]) != key
                    ]
                seen_comp[key] = True
                component_rows.append({
                    "sheet": sheet_name,
                    "excel_row": r_i,
                    "page_path": page_path.rstrip("/"),
                    "resourceType": rt,
                    "instance": instance,
                    "properties": props,
                    "component_label": sheet_name,
                })

    return {
        "status": "success",
        "page_rows": page_rows,
        "component_rows": component_rows,
        "duplicates_removed": duplicates_removed,
        "parse_errors": parse_errors,
        "summary": {
            "page_property_rows": len(page_rows),
            "component_rows": len(component_rows),
            "duplicates_removed": duplicates_removed,
            "parse_errors": len(parse_errors),
        },
    }


def _find_component_instance(
    aem: AEMClient, page_path: str, resource_type: str, instance: int
) -> Tuple[Optional[str], List[dict], Optional[str]]:
    """
    Find Nth (1-based) component on page matching resourceType (suffix or full).
    Accepts friendly names (Title, button) by resolving via dictionary.
    Returns (path, matching_list, error).
    """
    result = aem.get_components(page_path)
    if result.get("status") != "success":
        return None, [], result.get("message") or "Could not list components"

    comps = result.get("components") or []
    # Resolve friendly name → resourceType when needed
    rt = resource_type or ""
    if rt and "/" not in rt:
        try:
            dictionary = load_dictionary()
            resolved = _resolve_resource_type_from_name(rt, dictionary)
            if resolved:
                rt = resolved
        except Exception:
            pass
    rt_l = _norm(rt)
    rt_suffix = rt_l.split("/")[-1]
    # also match "title" leaf against core/.../title and weretail/.../title
    alt_suffixes = {rt_suffix}
    if rt_suffix in ("title", "heroimage", "hero_image", "button", "teaser"):
        alt_suffixes.add(rt_suffix.replace("_", ""))

    matching = []
    for c in comps:
        crt = _norm(c.get("resourceType") or "")
        leaf = crt.split("/")[-1]
        if (
            crt == rt_l
            or crt.endswith("/" + rt_suffix)
            or leaf == rt_suffix
            or leaf in alt_suffixes
            or rt_suffix in leaf
        ):
            matching.append(c)
    matching = sorted(matching, key=lambda x: x.get("path") or "")

    idx = instance - 1
    if idx < 0 or idx >= len(matching):
        return None, matching, (
            f"Component '{resource_type}' instance {instance} not found. "
            f"Found {len(matching)} instance(s)."
        )
    return matching[idx].get("path"), matching, None


def validate_all(
    parsed: dict,
    aem: Optional[AEMClient] = None,
    will_exist_paths: Optional[set] = None,
) -> dict:
    """Validate every row against AEM before any write.
    will_exist_paths: page paths that will be created in the same bulk batch (Pages sheet).
    SEO on those pages is planned update after create — not an error.
    """
    aem = aem or AEMClient()
    will_exist_paths = will_exist_paths or set()
    will_exist_paths = {p.rstrip("/") for p in will_exist_paths if p}
    page_results = []
    comp_results = []

    def _page_exists(page_path: str) -> bool:
        try:
            r = aem.session.get(
                f"{aem.base_url}{page_path}.json",
                timeout=min(getattr(aem, "timeout", 15) or 15, 10),
            )
            return r.status_code == 200
        except Exception:
            return False

    for row in parsed.get("page_rows") or []:
        page_path = (row["page_path"] or "").rstrip("/")
        target = f"{page_path}/jcr:content"
        item = {
            "type": "page_properties",
            "page_path": page_path,
            "target_path": target,
            "sheet": row.get("sheet"),
            "excel_row": row.get("excel_row"),
            "properties": row.get("properties") or {},
            "errors": [],
            "warnings": [],
            "validation": None,
            "page_status": None,
        }
        exists = _page_exists(page_path)
        if exists:
            item["page_status"] = "exists"
            # Fast property check — avoid heavy dialog discovery on preview
            try:
                validation = aem.validate_properties_for_update(
                    target, row.get("properties") or {}
                )
                item["validation"] = validation
                if validation.get("rejected"):
                    item["errors"].append(
                        "Rejected fields: "
                        + ", ".join(
                            f"{r.get('name')} ({r.get('reason')})"
                            for r in validation["rejected"]
                        )
                    )
                if validation.get("allowed"):
                    item["will_update"] = validation["allowed"]
            except Exception as e:
                # Still allow update attempt on apply
                item["warnings"].append(f"Could not pre-validate dialog fields: {e}")
                item["will_update"] = list((row.get("properties") or {}).keys())
        elif page_path in will_exist_paths:
            item["page_status"] = "will_create_in_batch"
            item["warnings"].append(
                "Page will be created in Pages step — SEO fields will be applied after create"
            )
            item["will_update"] = list((row.get("properties") or {}).keys())
        else:
            item["page_status"] = "missing"
            item["errors"].append(
                f"Page not found and not scheduled for create in Pages sheet: {page_path}"
            )
        page_results.append(item)

    for row in parsed.get("component_rows") or []:
        page_path = row["page_path"]
        rt = row["resourceType"]
        instance = row.get("instance") or 1
        item = {
            "type": "component",
            "page_path": page_path,
            "resourceType": rt,
            "instance": instance,
            "sheet": row.get("sheet"),
            "excel_row": row.get("excel_row"),
            "properties": row.get("properties") or {},
            "errors": [],
            "validation": None,
            "target_path": None,
        }
        path, matching, err = _find_component_instance(aem, page_path, rt, instance)
        if err:
            item["errors"].append(err)
            comp_results.append(item)
            continue
        item["target_path"] = path
        validation = aem.validate_properties_for_update(path, row.get("properties") or {})
        item["validation"] = validation
        if validation.get("rejected"):
            item["errors"].append(
                "Rejected fields: "
                + ", ".join(
                    f"{r.get('name')} ({r.get('reason')})"
                    for r in validation["rejected"]
                )
            )
        comp_results.append(item)

    error_count = sum(1 for x in page_results + comp_results if x.get("errors"))
    return {
        "status": "success",
        "page_results": page_results,
        "component_results": comp_results,
        "summary": {
            "total_rows": len(page_results) + len(comp_results),
            "rows_with_errors": error_count,
            "duplicates_removed": parsed.get("duplicates_removed", 0),
            "parse_errors": parsed.get("parse_errors") or [],
        },
    }


def apply_all(parsed: dict, performed_by: str = "system", aem: Optional[AEMClient] = None) -> dict:
    """
    Validate-aware apply:
      - Page properties → update jcr:content
      - Components → resolve instance → update
      - Applies valid fields even if some rejected (reported)
    """
    aem = aem or AEMClient()
    # Always re-validate structure via validate path resolution
    results = {
        "page_results": [],
        "component_results": [],
        "success_count": 0,
        "error_count": 0,
        "partial_count": 0,
        "skipped_count": 0,
    }

    for row in parsed.get("page_rows") or []:
        page_path = row["page_path"]
        target = f"{page_path}/jcr:content"
        row_out = {
            "page_path": page_path,
            "target_path": target,
            "sheet": row.get("sheet"),
            "errors": [],
            "updated_fields": [],
            "rejected": [],
            "skipped": [],
        }
        upd = aem.update_component(target, normalize_props(row.get("properties") or {}), performed_by=performed_by)
        if upd.get("status") in ("success", "partial"):
            row_out["updated_fields"] = upd.get("updated_properties") or []
            row_out["rejected"] = upd.get("rejected") or []
            row_out["skipped"] = upd.get("skipped") or []
            row_out["message"] = upd.get("message")
            if upd.get("status") == "partial" or row_out["rejected"]:
                results["partial_count"] += 1
            else:
                results["success_count"] += 1
            if not row_out["updated_fields"] and not row_out["rejected"]:
                results["skipped_count"] += 1
        else:
            row_out["errors"].append(upd.get("message") or "Update failed")
            row_out["rejected"] = upd.get("rejected") or []
            results["error_count"] += 1
        results["page_results"].append(row_out)

    for row in parsed.get("component_rows") or []:
        page_path = row["page_path"]
        rt = row["resourceType"]
        instance = row.get("instance") or 1
        row_out = {
            "page_path": page_path,
            "resourceType": rt,
            "instance": instance,
            "sheet": row.get("sheet"),
            "errors": [],
            "updated_fields": [],
            "rejected": [],
            "skipped": [],
        }
        path, matching, err = _find_component_instance(aem, page_path, rt, instance)
        if err:
            row_out["errors"].append(err)
            results["error_count"] += 1
            results["component_results"].append(row_out)
            continue
        row_out["target_path"] = path
        upd = aem.update_component(path, normalize_props(row.get("properties") or {}), performed_by=performed_by)
        if upd.get("status") in ("success", "partial"):
            row_out["updated_fields"] = upd.get("updated_properties") or []
            row_out["rejected"] = upd.get("rejected") or []
            row_out["skipped"] = upd.get("skipped") or []
            row_out["message"] = upd.get("message")
            if upd.get("status") == "partial" or row_out["rejected"]:
                results["partial_count"] += 1
            elif row_out["updated_fields"]:
                results["success_count"] += 1
            else:
                results["skipped_count"] += 1
        else:
            row_out["errors"].append(upd.get("message") or "Update failed")
            row_out["rejected"] = upd.get("rejected") or []
            results["error_count"] += 1
        results["component_results"].append(row_out)

    return {
        "status": "success",
        "message": (
            f"Completed – Success: {results['success_count']}, "
            f"Partial: {results['partial_count']}, "
            f"Errors: {results['error_count']}, "
            f"Skipped: {results['skipped_count']}"
        ),
        "results": results,
    }



def preview_excel(content: bytes, will_exist_paths: Optional[set] = None) -> dict:
    """
    Parse + validate-all. Response shape matches frontend:
      summary.total_seo_rows, summary.total_component_rows
      seo_updates[], component_updates[]
    will_exist_paths: pages that will be created in same bulk (from Pages sheet).
    """
    parsed = parse_workbook(content)
    if parsed.get("status") != "success":
        return parsed

    # Derive will-exist from Pages sheet if not provided
    if will_exist_paths is None:
        will_exist_paths = set()
        try:
            from backend.app.services.page_bulk_service import preview_pages
            pages = preview_pages(content)
            for plan in pages.get("plans") or []:
                p = plan.get("page_path")
                if not p:
                    continue
                if plan.get("action") in ("create_page", "exists", "will_create"):
                    will_exist_paths.add(p.rstrip("/"))
        except Exception:
            pass

    aem = AEMClient()
    validated = validate_all(parsed, aem=aem, will_exist_paths=will_exist_paths)

    seo_updates = []
    for row, vrow in zip(parsed.get("page_rows") or [], validated.get("page_results") or []):
        item = {
            "page_path": row["page_path"],
            "properties": row.get("properties") or {},
            "target_path": f"{row['page_path']}/jcr:content",
            "errors": vrow.get("errors") or [],
            "warnings": vrow.get("warnings") or [],
            "page_status": vrow.get("page_status"),
        }
        val = vrow.get("validation") or {}
        if val.get("rejected"):
            item["rejected_fields"] = val["rejected"]
        if val.get("allowed"):
            item["will_update"] = val["allowed"]
        elif vrow.get("will_update"):
            item["will_update"] = vrow.get("will_update")
        if val.get("skipped"):
            item["will_skip"] = val["skipped"]
        seo_updates.append(item)

    # if lengths mismatch (shouldn't), build from page_rows only
    if len(seo_updates) != len(parsed.get("page_rows") or []):
        seo_updates = [
            {
                "page_path": r["page_path"],
                "properties": r.get("properties") or {},
                "target_path": f"{r['page_path']}/jcr:content",
            }
            for r in (parsed.get("page_rows") or [])
        ]

    component_updates = []
    for row, vrow in zip(parsed.get("component_rows") or [], validated.get("component_results") or []):
        item = {
            "page_path": row["page_path"],
            "component_name": row.get("component_label") or row.get("resourceType") or "",
            "resourceType": row.get("resourceType"),
            "instance": row.get("instance") or 1,
            "properties": row.get("properties") or {},
            "component_path": vrow.get("target_path"),
            "errors": vrow.get("errors") or [],
        }
        val = vrow.get("validation") or {}
        if val.get("rejected"):
            item["rejected_fields"] = val["rejected"]
        if val.get("allowed"):
            item["will_update"] = val["allowed"]
        if val.get("skipped"):
            item["will_skip"] = val["skipped"]
        component_updates.append(item)

    if len(component_updates) != len(parsed.get("component_rows") or []):
        component_updates = [
            {
                "page_path": r["page_path"],
                "component_name": r.get("component_label") or r.get("resourceType") or "",
                "resourceType": r.get("resourceType"),
                "instance": r.get("instance") or 1,
                "properties": r.get("properties") or {},
            }
            for r in (parsed.get("component_rows") or [])
        ]

    sheets = []
    for r in (parsed.get("page_rows") or []):
        if r.get("sheet") and r["sheet"] not in sheets:
            sheets.append(r["sheet"])
    for r in (parsed.get("component_rows") or []):
        if r.get("sheet") and r["sheet"] not in sheets:
            sheets.append(r["sheet"])

    return {
        "status": "success",
        "message": "Preview generated successfully",
        "summary": {
            "total_seo_rows": len(seo_updates),
            "total_component_rows": len(component_updates),
            "sheets_processed": sheets,
            "duplicates_removed": parsed.get("duplicates_removed", 0),
            "parse_errors": parsed.get("parse_errors") or [],
            "rows_with_errors": (validated.get("summary") or {}).get("rows_with_errors", 0),
        },
        "seo_updates": seo_updates,
        "component_updates": component_updates,
        "validation": validated,
    }


def apply_excel(content: bytes, performed_by: str = "system") -> dict:
    """
    Apply all rows. Response shape matches frontend:
      results.seo_results, results.component_results
      success_count, error_count, skipped_count
    """
    parsed = parse_workbook(content)
    if parsed.get("status") != "success":
        return parsed

    aem = AEMClient()
    raw = apply_all(parsed, performed_by=performed_by, aem=aem)
    res = raw.get("results") or {}

    # Map to frontend keys
    seo_results = []
    for r in res.get("page_results") or []:
        skipped = r.get("skipped") or []
        skipped_fields = []
        for s in skipped:
            if isinstance(s, dict):
                skipped_fields.append(s.get("name") or s.get("message") or str(s))
            else:
                skipped_fields.append(str(s))
        rejected = r.get("rejected") or []
        errors = list(r.get("errors") or [])
        for rej in rejected:
            if isinstance(rej, dict):
                errors.append(f"{rej.get('name')}: {rej.get('reason') or rej.get('message')}")
            else:
                errors.append(str(rej))
        seo_results.append({
            "page_path": r.get("page_path"),
            "component_path": r.get("target_path"),
            "updated_fields": r.get("updated_fields") or [],
            "skipped_fields": skipped_fields,
            "errors": errors,
            "message": r.get("message"),
        })

    component_results = []
    for r in res.get("component_results") or []:
        skipped = r.get("skipped") or []
        skipped_fields = []
        for s in skipped:
            if isinstance(s, dict):
                skipped_fields.append(s.get("name") or s.get("message") or str(s))
            else:
                skipped_fields.append(str(s))
        rejected = r.get("rejected") or []
        errors = list(r.get("errors") or [])
        for rej in rejected:
            if isinstance(rej, dict):
                errors.append(f"{rej.get('name')}: {rej.get('reason') or rej.get('message')}")
            else:
                errors.append(str(rej))
        component_results.append({
            "page_path": r.get("page_path"),
            "component_name": r.get("resourceType") or r.get("sheet") or "component",
            "instance": r.get("instance") or 1,
            "component_path": r.get("target_path"),
            "updated_fields": r.get("updated_fields") or [],
            "skipped_fields": skipped_fields,
            "errors": errors,
            "message": r.get("message"),
        })

    return {
        "status": "success",
        "message": raw.get("message") or "Apply completed",
        "results": {
            "seo_results": seo_results,
            "component_results": component_results,
            "success_count": res.get("success_count", 0),
            "error_count": res.get("error_count", 0),
            "skipped_count": res.get("skipped_count", 0),
            "partial_count": res.get("partial_count", 0),
        },
    }


# =============================================================================
# Full bulk orchestration: Assets → Pages → Add components → Update
# =============================================================================

def _is_add_sheet(name: str) -> bool:
    return (name or "").strip().lower().startswith("add ")


def parse_add_sheets(content: bytes) -> dict:
    """Parse 'Add *' sheets into component-add rows (CA component name + properties)."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    dictionary = load_dictionary()
    rows = []
    errors = []

    for sheet_name in wb.sheetnames:
        if not _is_add_sheet(sheet_name):
            continue
        # "Add Hero Image" → default component name Hero Image
        default_name = sheet_name.strip()[4:].strip()
        ws = wb[sheet_name]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        headers = [str(h).strip() if h is not None else "" for h in data[0]]
        header_map = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            hl = _norm(h)
            if hl in ("page path", "pagepath", "path"):
                header_map[idx] = "__page_path__"
            elif hl in ("component name", "component", "name"):
                header_map[idx] = "__component_name__"
            else:
                # resolve via dictionary using default component leaf
                resolved = None
                for rt, meta in (dictionary or {}).items():
                    if not isinstance(meta, dict):
                        continue
                    if _norm(meta.get("label") or "") == _norm(default_name) or rt.split("/")[-1].lower() == _norm(default_name).replace(" ", ""):
                        resolved = resolve_label(rt, h) or None
                        if not resolved:
                            fields = meta.get("fields") or {}
                            for fn, aliases in fields.items():
                                alist = aliases if isinstance(aliases, list) else [str(aliases)]
                                if _norm(fn) == hl or any(_norm(a) == hl for a in alist):
                                    resolved = fn
                                    break
                        if resolved:
                            break
                resolved = resolved or h
                # Common CA labels for image fields
                rl = _norm(str(resolved))
                if rl in ("image", "file", "file reference", "asset", "hero image"):
                    resolved = "fileReference"
                elif rl in ("full width", "usefullwidth"):
                    resolved = "useFullWidth"
                elif rl in ("button label", "buttonlabel"):
                    resolved = "buttonLabel"
                elif rl in ("button link", "buttonlink", "button link to"):
                    resolved = "buttonLinkTo"
                header_map[idx] = resolved

        for r_i, row in enumerate(data[1:], start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            page_path = None
            comp_name = default_name
            props = {}
            for idx, key in header_map.items():
                if idx >= len(row):
                    continue
                val = row[idx]
                if val is None or str(val).strip() == "":
                    continue  # skip blank cells
                if key == "__page_path__":
                    page_path = str(val).strip()
                elif key == "__component_name__":
                    comp_name = str(val).strip()
                else:
                    props[key] = val if not isinstance(val, str) else val.strip()
            if not page_path:
                errors.append(f"{sheet_name} row {r_i}: missing Page Path")
                continue
            rows.append({
                "sheet": sheet_name,
                "excel_row": r_i,
                "page_path": page_path.rstrip("/"),
                "component": comp_name,
                "properties": props,
            })

    return {"status": "success", "rows": rows, "errors": errors}


def orchestrate_preview(content: bytes, username: str = None, session_id: str = None) -> dict:
    """
    Preview full bulk workbook in order (no writes).
    Validates Add-sheet page paths against existing AEM pages + Pages sheet Create=Y.
    """
    from backend.app.services.asset_bulk_service import plan_asset_uploads
    from backend.app.services.page_bulk_service import preview_pages
    from backend.app.services.page_service import PageService
    from backend.app.services.component_add_service import ComponentAddService

    assets = plan_asset_uploads(content)
    pages = preview_pages(content)
    adds = parse_add_sheets(content)
    # Build will_exist early so SEO preview knows pages created in this batch
    will_exist_early = set()
    for plan in pages.get("plans") or []:
        p = plan.get("page_path")
        if p and plan.get("action") in ("create_page", "exists", "will_create"):
            will_exist_early.add(p.rstrip("/"))
    updates = preview_excel(content, will_exist_paths=will_exist_early)

    ps = PageService()
    add_svc = ComponentAddService()

    # Paths that will exist after Pages step
    will_exist = set()
    for plan in pages.get("plans") or []:
        p = plan.get("page_path")
        if not p:
            continue
        if plan.get("action") in ("create_page", "exists"):
            will_exist.add(p.rstrip("/"))
        if plan.get("action") == "exists":
            will_exist.add(p.rstrip("/"))

    # Also any path that already exists in AEM
    validated_rows = []
    add_errors = list(adds.get("errors") or [])
    add_warnings = []
    for row in adds.get("rows") or []:
        entry = dict(row)
        entry["errors"] = []
        entry["warnings"] = []
        page_path = (row.get("page_path") or "").rstrip("/")
        exists_now = False
        try:
            exists_now = ps.path_exists(page_path)
        except Exception:
            exists_now = False
        if exists_now:
            entry["page_status"] = "exists"
            will_exist.add(page_path)
        elif page_path in will_exist:
            entry["page_status"] = "will_create_in_batch"
            entry["warnings"].append("Page will be created in Pages step of this same bulk run")
        else:
            entry["page_status"] = "missing"
            entry["errors"].append(
                f"Page path not found and not scheduled for create in Pages sheet: {page_path}"
            )
            add_errors.append(f"{row.get('sheet')} row {row.get('excel_row')}: page missing: {page_path}")

        # Resolve component name
        comp = row.get("component") or ""
        res = add_svc.resolve_resource_type(comp, page_path if exists_now else None)
        if res.get("status") != "success":
            entry["errors"].append(res.get("message") or f"Cannot resolve component '{comp}'")
            add_errors.append(
                f"{row.get('sheet')} row {row.get('excel_row')}: cannot resolve '{comp}'"
            )
        else:
            entry["resourceType"] = res.get("resourceType")
            entry["resolved_by"] = res.get("matched_by")
            # Allowed check when page exists
            if exists_now or page_path in will_exist:
                try:
                    allowed = add_svc.get_allowed_components(page_path if exists_now else page_path)
                    allowed_set = set(allowed.get("allowed_resource_types") or [])
                    rt = res.get("resourceType") or ""
                    if allowed_set and rt not in allowed_set:
                        leaf = rt.split("/")[-1]
                        if not any(leaf == a.split("/")[-1] for a in allowed_set):
                            entry["warnings"].append(
                                f"Component may not be in allowed list for this layout ({rt})"
                            )
                            add_warnings.append(entry["warnings"][-1])
                except Exception:
                    pass

        # Empty properties warning
        if not (row.get("properties") or {}):
            entry["warnings"].append("No field values provided — component will be added with defaults only")

        # Explicit action for UI — Add sheet always means NEW instance (page may already exist)
        if entry.get("errors"):
            entry["action"] = "blocked"
        else:
            entry["action"] = "add_new_instance"
            entry["action_label"] = "Will ADD new component instance on page"

        validated_rows.append(entry)

    blocked = sum(1 for r in validated_rows if r.get("errors"))
    ok = sum(1 for r in validated_rows if not r.get("errors"))


    # --- Recommendations: page paths vs DAM asset targets (not hard-coded projects) ---
    recommendations = []
    page_paths_for_dam = []
    for plan in pages.get("plans") or []:
        if plan.get("action") in ("create_page", "exists") and plan.get("page_path"):
            page_paths_for_dam.append(plan["page_path"].rstrip("/"))

    asset_targets = []
    for p in assets.get("plans") or []:
        if p.get("target_path"):
            asset_targets.append({
                "target": p["target_path"].rstrip("/"),
                "source": p.get("source_path"),
                "row": p.get("excel_row"),
                "errors": p.get("errors") or [],
            })

    def suggest_dam_for_page(page_path: str) -> list:
        """
        Dynamic suggestions from page path only (no project hardcoding).
        /content/<site>/.../<page> → /content/dam/<site>/.../<page> variants
        """
        parts = [x for x in page_path.split("/") if x]
        if len(parts) < 2 or parts[0] != "content":
            return []
        after = parts[1:]  # site, ...
        suggestions = []
        # Full mirror under /content/dam
        suggestions.append("/content/dam/" + "/".join(after))
        # Drop locale-ish middle segments if 4+ (keep site + last 2)
        if len(after) >= 3:
            suggestions.append("/content/dam/" + "/".join([after[0]] + after[-2:]))
        if len(after) >= 2:
            suggestions.append("/content/dam/" + "/".join([after[0], after[-1]]))
        # unique
        out = []
        for s in suggestions:
            if s not in out:
                out.append(s)
        return out

    for pp in page_paths_for_dam:
        leaf = pp.split("/")[-1]
        suggestions = suggest_dam_for_page(pp)
        # Does any asset target equal suggestion or end with /leaf (and look page-specific)?
        matched = False
        for at in asset_targets:
            t = at["target"]
            if t in suggestions or t.rstrip("/").endswith("/" + leaf):
                # weak match on leaf only if target deeper than dam root+1
                if t in suggestions or t.count("/") >= 5:
                    matched = True
                    break
        if not matched and suggestions:
            recommendations.append({
                "type": "dam_path_alignment",
                "severity": "warn",
                "page_path": pp,
                "message": (
                    f"Page '{pp}' has no Assets Target Path that looks aligned. "
                    f"Recommended DAM folder(s) for this page’s assets: {', '.join(suggestions)}. "
                    "Update the Assets sheet and re-upload Excel if this page needs its own assets."
                ),
                "suggested_dam_paths": suggestions,
            })

    for at in asset_targets:
        if at["errors"]:
            continue
        t = at["target"]
        # If target leaf doesn't match any page leaf being created, warn
        tleaf = t.split("/")[-1]
        page_leaves = {p.split("/")[-1] for p in page_paths_for_dam}
        if page_paths_for_dam and tleaf not in page_leaves and tleaf in ("men", "women", "en", "us", "ca"):
            recommendations.append({
                "type": "dam_target_generic",
                "severity": "warn",
                "excel_row": at["row"],
                "target_path": t,
                "message": (
                    f"Assets row {at['row']} targets '{t}' which looks like a parent/section folder, "
                    f"while Pages create: {', '.join(page_paths_for_dam)}. "
                    "For page-specific assets, Target Path usually ends with the page name "
                    "(e.g. .../men/test or .../men/test/test-1)."
                ),
            })

    # Any component field that points at DAM (image, pdf, video, json, etc.) — dynamic
    # Detection is primarily by VALUE (/content/dam/...), not by component type.
    ASSET_FIELD_HINTS = (
        "image", "file", "asset", "thumbnail", "poster", "video", "pdf",
        "document", "json", "media", "attachment", "download", "fileref",
        "filereference", "file reference", "dam", "rendition",
    )

    def looks_like_dam_path(val: str) -> bool:
        v = (val or "").strip().replace("\\", "/")
        return v.startswith("/content/dam/") or v.startswith("content/dam/")

    def is_likely_asset_field(name: str) -> bool:
        n = _norm(name or "")
        if not n:
            return False
        for h in ASSET_FIELD_HINTS:
            if h in n:
                return True
        return False

    # Map page -> list of asset targets from Excel for that page leaf
    targets_by_leaf = {}
    for at in asset_targets:
        leaf = at["target"].split("/")[-1]
        targets_by_leaf.setdefault(leaf, []).append(at["target"])

    for row in validated_rows:
        page_path = (row.get("page_path") or "").rstrip("/")
        page_leaf = page_path.split("/")[-1] if page_path else ""
        props = row.get("properties") or {}
        suggested = suggest_dam_for_page(page_path) if page_path else []
        page_targets = targets_by_leaf.get(page_leaf) or []
        for fk, fval in props.items():
            aval = str(fval or "").strip()
            if not aval:
                continue
            # Dynamic: any DAM path value, on any component — not Hero-only
            if not looks_like_dam_path(aval) and not is_likely_asset_field(str(fk)):
                continue
            if not looks_like_dam_path(aval):
                continue
            if not aval.startswith("/"):
                aval = "/" + aval
            aligned = False
            # Align if asset path is under suggested DAM or under any Assets target for this page leaf
            for s in suggested:
                if aval == s or aval.startswith(s.rstrip("/") + "/"):
                    aligned = True
                    break
            if not aligned:
                for t in page_targets:
                    if aval == t or aval.startswith(t.rstrip("/") + "/"):
                        aligned = True
                        break
            if not aligned:
                # also OK if aval contains page leaf as path segment
                if f"/{page_leaf}/" in aval or aval.endswith("/" + page_leaf):
                    aligned = True
            if not aligned:
                msg = (
                    f"Component '{row.get('component')}' on page '{page_path}' uses asset path "
                    f"'{aval}' which does not align with recommended DAM folder(s) for this page"
                )
                if suggested:
                    msg += f": {', '.join(suggested)}"
                msg += ". Update this DAM path in the Add sheet to match the Assets Target Path for this page (project recommendation). Applies to any asset field: image, PDF, video, JSON, etc."
                recommendations.append({
                    "type": "component_asset_alignment",
                    "severity": "warn",
                    "page_path": page_path,
                    "component": row.get("component"),
                    "field": fk,
                    "asset_path": aval,
                    "suggested_dam_paths": suggested,
                    "excel_row": row.get("excel_row"),
                    "sheet": row.get("sheet"),
                    "message": msg,
                })
                row.setdefault("warnings", []).append(msg)

    session_delta = None
    try:
        from backend.app.services.bulk_session_service import diff_against_session
        if session_id:
            session_delta = diff_against_session(session_id, content)
    except Exception as e:
        session_delta = {"status": "error", "message": str(e)}

    return {
        "status": "success",
        "message": "Full bulk preview (Assets → Pages → Add → Update)",
        "assets": assets,
        "pages": pages,
        "components_add": {
            "status": "success" if blocked == 0 else "partial",
            "rows": validated_rows,
            "errors": add_errors,
            "warnings": add_warnings,
            "summary": {
                "total": len(validated_rows),
                "ok": ok,
                "blocked": blocked,
            },
        },
        "updates": updates,
        "recommendations": recommendations,
        "session_delta": session_delta,
    }


def orchestrate_apply(content: bytes, performed_by: str = "system") -> dict:
    """
    Apply full bulk workbook in order.
    Continues past per-row failures; blank fields already skipped in parsers.
    """
    from backend.app.services.asset_bulk_service import plan_asset_uploads
    from backend.app.services.page_bulk_service import apply_pages, preview_pages
    from backend.app.services.dam_service import DamService
    from backend.app.services.component_add_service import ComponentAddService
    from collections import OrderedDict

    results = {
        "assets": None,
        "pages": None,
        "components_add": None,
        "updates": None,
    }

    # 1) Assets
    try:
        plan = plan_asset_uploads(content)
        dam = DamService()
        asset_results = []
        if plan.get("status") == "success":
            for p in plan.get("plans") or []:
                if p.get("errors"):
                    asset_results.append({"row": p.get("excel_row"), "status": "error", "errors": p["errors"]})
                    continue
                up = dam.upload_from_local(
                    page_dam_path=p["target_path"],
                    local_page_folder=p["source_path"],
                    confirm_create_folders=True,
                )
                asset_results.append({"row": p.get("excel_row"), "upload": up})
            results["assets"] = {"status": "success", "results": asset_results, "plan_summary": plan.get("summary")}
        else:
            results["assets"] = plan
    except Exception as e:
        results["assets"] = {"status": "error", "message": str(e)}

    # 2) Pages
    try:
        results["pages"] = apply_pages(content, performed_by=performed_by)
    except Exception as e:
        results["pages"] = {"status": "error", "message": str(e)}

    # 3) Component ADD — only on pages that exist; skip others
    try:
        from backend.app.services.page_service import PageService
        adds = parse_add_sheets(content)
        by_page = OrderedDict()
        for row in adds.get("rows") or []:
            by_page.setdefault(row["page_path"], []).append(row)
        add_svc = ComponentAddService()
        ps = PageService()
        add_results = []
        for page_path, rows in by_page.items():
            if not ps.path_exists(page_path):
                add_results.append({
                    "page_path": page_path,
                    "status": "skipped",
                    "message": "Page does not exist — skipped component add (create page first)",
                    "components_requested": [r.get("component") for r in rows],
                })
                continue
            components = [
                {"component": r["component"], "properties": r.get("properties") or {}}
                for r in rows
            ]
            r = add_svc.apply_add(page_path, components)
            add_results.append({"page_path": page_path, "status": r.get("status"), "result": r})
        any_ok = any(x.get("status") == "success" for x in add_results)
        any_err = any(x.get("status") not in ("success", "skipped") for x in add_results)
        results["components_add"] = {
            "status": "success" if any_ok and not any_err else ("partial" if any_ok else "error"),
            "pages": add_results,
            "parse_errors": adds.get("errors") or [],
        }
    except Exception as e:
        results["components_add"] = {"status": "error", "message": str(e)}

    # 4) SEO + component UPDATE
    try:
        results["updates"] = apply_excel(content, performed_by=performed_by)
    except Exception as e:
        results["updates"] = {"status": "error", "message": str(e)}

    return {
        "status": "success",
        "message": "Full bulk apply finished (see each stage for details)",
        "stages": results,
    }
