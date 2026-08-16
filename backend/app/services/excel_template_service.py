"""
Bulk Excel template generator — always returns valid xlsx bytes or raises.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _preferred_label(aliases) -> str:
    if isinstance(aliases, list) and aliases:
        return str(aliases[0])
    if isinstance(aliases, str) and aliases:
        return aliases.split(",")[0].strip()
    return ""


def _load_dict_safe() -> dict:
    try:
        from backend.app.services.dictionary_service import load_dictionary
        return load_dictionary() or {}
    except Exception as e:
        logger.warning("dictionary load failed: %s", e)
        return {}


def _style_header(ws, headers: list, fill: PatternFill, font: Font, border: Border):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, str(header) if header is not None else "")
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        cell.protection = Protection(locked=True)
    for col in range(1, max(len(headers), 1) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 32 if col > 1 else 48


def _empty_rows(ws, headers: list, border: Border, n: int = 8):
    for row in range(2, 2 + n):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col, "")
            cell.border = border
            cell.protection = Protection(locked=False)


def _unique_sheet_name(wb: Workbook, base: str) -> str:
    name = (base or "Sheet")[:31]
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        candidate = f"{base[:28]}_{i}"[:31]
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def _field_header(data: dict, resource_type: str, field_name: str) -> str:
    try:
        entry = data.get(resource_type) if isinstance(data, dict) else None
        if isinstance(entry, dict):
            fields = entry.get("fields") or {}
            if isinstance(fields, dict):
                label = _preferred_label(fields.get(field_name))
                if label:
                    return label
    except Exception:
        pass
    return field_name


def generate_template(
    selections: List[dict],
    include_seo: bool = True,
    include_assets: bool = True,
    include_pages: bool = True,
    include_components_add: bool = True,
    include_components_update: bool = False,
    known_templates: Optional[List[dict]] = None,
    allowed_components: Optional[List[dict]] = None,
    default_template_name: Optional[str] = None,
) -> bytes:
    data = _load_dict_safe()
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    example_font = Font(color="94A3B8", italic=True)

    # 1. Instructions
    ws0 = wb.active
    ws0.title = "Instructions"
    lines = [
        "AEM CONTENT UPDATER — BULK EXCEL TEMPLATE",
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC",
        "",
        "SHEET ORDER",
        "1. Assets — upload images to DAM",
        "2. Pages — create new pages",
        "3. Add Title / Add Hero Image / ... — ADD new components on a page (top to bottom = order on page)",
        "4. Page Properties SEO — page-level fields",
        "5. Title / Hero Image / ... — UPDATE existing components already on the page",
        "",
        "ENABLED TEMPLATES (use these names in the Pages sheet → Template Name column)",
    ]
    for t in known_templates or []:
        try:
            if t.get("jcr:primaryType") and t.get("jcr:primaryType") != "cq:Template":
                continue
            title = (t.get("title") or t.get("name") or "").strip()
            if title:
                lines.append(f"  • {title}")
        except Exception:
            continue
    if not known_templates:
        lines.append("  • Content Page")

    lines.extend(["", "ALLOWED COMPONENT NAMES (use these in Add sheets → Component Name column)"])
    for c in allowed_components or []:
        try:
            n = (c.get("name") or "").strip()
            if n:
                lines.append(f"  • {n}")
        except Exception:
            continue
    if not allowed_components:
        lines.append("  • Title")
        lines.append("  • Hero Image")

    lines.extend([
        "",
        "Do not rename header rows. Always Preview before Apply.",
        "See sheet: How to use",
    ])
    for i, line in enumerate(lines, 1):
        ws0.cell(i, 1, line)
        if i == 1:
            ws0.cell(i, 1).font = Font(bold=True, size=14, color="1E3A5F")
    ws0.column_dimensions["A"].width = 110

    # How to use
    ws_h = wb.create_sheet("How to use", 1)
    how_lines = [
        "HOW TO USE THIS WORKBOOK",
        "",
        "ASSETS sheet",
        "  Source Path = folder on your PC that contains Desktop, Mobile, Tablet subfolders",
        "  Target Path = DAM folder (example: /content/dam/we-retail/en/men)",
        "  One row per page asset folder. The tool finds breakpoint folders and uploads files.",
        "",
        "PAGES sheet",
        "  Page Path = full path of the page to create",
        "  Create (Y/N) = Y to create if missing",
        "  Template Name = pick a simple name from Instructions (example: Content Page)",
        "",
        "ADD sheets (example: Add Title, Add Hero Image)",
        "  Use these when the component is NOT yet on the page (new page or empty layout).",
        "  Page Path = page where component should be added",
        "  Component Name = simple name from Instructions (example: Title)",
        "  Other columns = field values for that component",
        "  Row order for the same Page Path = order components appear on the page (first row = top).",
        "",
        "UPDATE sheets (example: Title, Hero Image — without the word Add)",
        "  Use these when the component ALREADY exists on the page.",
        "  Page Path + Instance (1 = first Title, 2 = second Title, ...)",
        "  Fill only fields you want to change; leave others blank.",
        "",
        "PAGE PROPERTIES SEO sheet",
        "  One row per page for title, description, and other page dialog fields.",
        "",
        "TIPS",
        "  • Do not change header row text",
        "  • Use Preview in the tool before Apply",
        "  • For Hero-like components with one image field, prefer the desktop DAM asset path",
    ]
    for i, line in enumerate(how_lines, 1):
        ws_h.cell(i, 1, line)
        if i == 1:
            ws_h.cell(i, 1).font = Font(bold=True, size=14, color="1E3A5F")
    ws_h.column_dimensions["A"].width = 100

    # 2. Assets
    if include_assets:
        ws = wb.create_sheet("Assets")
        headers = ["Source Path", "Target Path"]
        _style_header(ws, headers, header_fill, header_font, thin)
        _empty_rows(ws, headers, thin, 10)
        ws.cell(2, 1, "C:/Users/you/assets/men")
        ws.cell(2, 2, "/content/dam/we-retail/en/men")
        for c in range(1, 3):
            ws.cell(2, c).font = example_font

    # 3. Pages
    if include_pages:
        ws = wb.create_sheet("Pages")
        headers = ["Page Path", "Create (Y/N)", "Template Name"]
        _style_header(ws, headers, header_fill, header_font, thin)
        _empty_rows(ws, headers, thin, 10)
        ws.cell(2, 1, "/content/we-retail/us/en/men/test")
        ws.cell(2, 2, "Y")
        default_tpl = (default_template_name or "").strip() or "Content Page"
        if not default_template_name:
            for t in known_templates or []:
                if (t.get("name") or "") == "content-page" or (t.get("title") or "").lower() == "content page":
                    default_tpl = t.get("title") or default_tpl
                    break
        ws.cell(2, 3, default_tpl)
        for c in range(1, 4):
            ws.cell(2, c).font = example_font

    selections = selections or []

    # 4. Page Properties SEO — only selected fields from page_properties selection
    page_prop_sel = None
    for sel in selections:
        if not isinstance(sel, dict):
            continue
        rt = (sel.get("resourceType") or "")
        label = (sel.get("label") or "").lower()
        if rt == "page_properties" or "page properties" in label or label == "seo":
            page_prop_sel = sel
            break

    if include_seo or page_prop_sel:
        headers = ["Page Path"]
        if page_prop_sel and page_prop_sel.get("fields"):
            for fn in page_prop_sel.get("fields") or []:
                headers.append(_field_header(data, "page_properties", str(fn)))
        else:
            # fallback minimal
            headers.extend(["Title", "Page Title", "Description"])
        ws = wb.create_sheet("Page Properties SEO")
        _style_header(ws, headers, header_fill, header_font, thin)
        _empty_rows(ws, headers, thin, 6)


    # 5. Components Add sheets — per selection include_add
    for sel in selections:
        if not isinstance(sel, dict):
            continue
        rt = sel.get("resourceType") or ""
        if rt == "page_properties" or "page properties" in (sel.get("label") or "").lower():
            continue
        want_add = sel.get("include_add")
        if want_add is None:
            want_add = bool(include_components_add)
        if include_pages and sel.get("include_add") is None:
            want_add = True
        if not want_add:
            continue
        label = str(sel.get("label") or (rt.split("/")[-1] if rt else "Component"))[:20]
        ws = wb.create_sheet(_unique_sheet_name(wb, f"Add {label}"))
        fields = sel.get("fields") or []
        headers = ["Page Path", "Component Name"]
        for fn in fields:
            headers.append(_field_header(data, rt, str(fn)))
        _style_header(ws, headers, header_fill, header_font, thin)
        _empty_rows(ws, headers, thin, 6)
        ws.cell(2, 1, "/content/we-retail/us/en/men/test")
        ws.cell(2, 2, label)
        for c in range(1, 3):
            ws.cell(2, c).font = example_font

    # 6. Component UPDATE sheets
    # Update sheets (Instance) — per selection include_update; never when include_pages
    if not include_pages:
        for sel in selections:
            if not isinstance(sel, dict):
                continue
            rt = sel.get("resourceType") or ""
            if rt == "page_properties" or "page properties" in (sel.get("label") or "").lower():
                continue
            want_upd = sel.get("include_update")
            if want_upd is None:
                want_upd = include_components_update
            if not want_upd:
                continue
            label = str(sel.get("label") or (rt.split("/")[-1] if rt else "Component"))[:28]
            ws = wb.create_sheet(_unique_sheet_name(wb, label))
            fields = sel.get("fields") or []
            headers = ["Page Path", "Instance"]
            for fn in fields:
                headers.append(_field_header(data, rt, str(fn)))
            _style_header(ws, headers, header_fill, header_font, thin)
            _empty_rows(ws, headers, thin, 6)

    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    if not raw.startswith(b"PK"):
        raise RuntimeError("Generated file is not a valid xlsx (missing ZIP header)")
    return raw


def describe_template_from_bytes(content: bytes) -> dict:
    """Inspect a generated workbook — same file CA downloads."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []):
            if cell is None:
                break
            headers.append(str(cell).strip())
        sheets.append({"name": name, "headers": headers})
    return {"status": "success", "sheet_count": len(sheets), "sheets": sheets}


def describe_template(
    selections: List[dict],
    include_seo: bool = True,
    include_assets: bool = True,
    include_pages: bool = True,
    include_components_add: bool = True,
    include_components_update: bool = False,
    known_templates: Optional[List[dict]] = None,
    allowed_components: Optional[List[dict]] = None,
    default_template_name: Optional[str] = None,
) -> dict:
    """
    Build the real template in memory, then list sheets/headers.
    Preview always matches download — no separate UI sheet list.
    """
    content = generate_template(
        selections,
        include_seo=include_seo,
        include_assets=include_assets,
        include_pages=include_pages,
        include_components_add=include_components_add,
        include_components_update=include_components_update,
        known_templates=known_templates,
        allowed_components=allowed_components,
        default_template_name=default_template_name,
    )
    info = describe_template_from_bytes(content)
    info["message"] = "Structure from the same generator as Create Template download"
    return info

