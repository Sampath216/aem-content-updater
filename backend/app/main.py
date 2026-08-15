"""
AEM Content Updater — FastAPI application entrypoint.
Clean single copy of each route. Do not register the same path twice.
"""
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import Request, Body, Depends, FastAPI, File, HTTPException, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from fastapi_mcp import FastApiMCP
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session
import io

from backend.app.core.config import get_settings
from backend.app.core.auth import (
    ACCESS_TOKEN_EXPIRE_MINUTES,
    authenticate_user,
    create_access_token,
    get_current_user,
    get_db,
    get_password_hash,
)
from backend.app.models.audit import AuditLog, SessionLocal
from backend.app.models.user import User
from backend.app.services.aem_client import AEMClient
from backend.app.services.asset_bulk_service import plan_asset_uploads
from backend.app.services.component_add_service import ComponentAddService
from backend.app.services.component_catalog import ComponentCatalog
from backend.app.services.dam_service import DamService
from backend.app.services.dictionary_service import (
    list_components,
    load_dictionary,
    sync_from_catalog_fields,
    update_field_aliases,
    upsert_component,
)
from backend.app.services.excel_bulk_service import apply_excel, preview_excel, orchestrate_preview, orchestrate_apply
from backend.app.services.bulk_validation_service import build_validation_report
from backend.app.services.bulk_session_service import mark_applied, clear_session, get_session
from backend.app.services.validation_export_service import validation_report_to_xlsx
from backend.app.services.excel_service import ExcelTemplateService
from backend.app.services.excel_template_service import generate_template
from backend.app.services.page_bulk_service import apply_pages, preview_pages
from backend.app.services.page_service import PageService
from backend.app.services.template_history_service import (
    delete_template,
    get_template,
    list_templates,
    mark_used,
    save_template,
)

settings = get_settings()

app = FastAPI(
    title=settings.APP_NAME,
    description="Enterprise tool to discover and update AEM components",
    version=settings.APP_VERSION,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =============================================================================
# AUTH
# =============================================================================

@app.post("/api/auth/login")
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = create_access_token(
        data={"sub": user.username},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "username": user.username,
        "full_name": user.full_name,
    }


@app.post("/api/auth/register")
def register_user(
    username: str = Body(...),
    password: str = Body(...),
    full_name: str = Body(None),
    db: Session = Depends(get_db),
):
    if db.query(User).filter(User.username == username).first():
        raise HTTPException(status_code=400, detail="Username already registered")
    user = User(
        username=username,
        hashed_password=get_password_hash(password),
        full_name=full_name or username,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {
        "status": "success",
        "message": f"User '{username}' created successfully",
        "username": user.username,
    }


# =============================================================================
# PUBLIC
# =============================================================================

@app.get("/")
def home():
    return {
        "message": f"{settings.APP_NAME} is running successfully!",
        "status": "healthy",
        "version": settings.APP_VERSION,
    }


@app.get("/health")
def health_check():
    return {"status": "ok", "app": settings.APP_NAME}


@app.get("/api/aem/status")
def aem_status():
    return AEMClient().is_reachable()


# =============================================================================
# AEM COMPONENTS
# =============================================================================

@app.get("/api/aem/components")
def get_page_components(page_path: str, current_user: User = Depends(get_current_user)):
    return AEMClient().get_components(page_path)


@app.get("/api/aem/component/fields")
def get_component_fields(component_path: str, current_user: User = Depends(get_current_user)):
    return AEMClient().get_component_fields(component_path)


@app.post("/api/aem/component/update")
def update_component(
    component_path: str,
    properties: Dict[str, Any] = Body(...),
    current_user: User = Depends(get_current_user),
):
    return AEMClient().update_component(
        component_path=component_path,
        properties=properties,
        performed_by=current_user.full_name or current_user.username,
    )


@app.get("/api/aem/component/diagnose")
def diagnose_component(component_path: str, current_user: User = Depends(get_current_user)):
    return AEMClient().diagnose_component_dialog(component_path)


@app.get("/api/audit/logs")
def get_audit_logs(limit: int = 50):
    db = SessionLocal()
    try:
        logs = (
            db.query(AuditLog)
            .order_by(AuditLog.timestamp.desc())
            .limit(limit)
            .all()
        )
        return {
            "status": "success",
            "count": len(logs),
            "logs": [
                {
                    "id": log.id,
                    "timestamp": log.timestamp.isoformat() if log.timestamp else None,
                    "component_path": log.component_path,
                    "property_name": log.property_name,
                    "old_value": log.old_value,
                    "new_value": log.new_value,
                    "success": log.success,
                    "message": log.message,
                    "performed_by": log.performed_by,
                }
                for log in logs
            ],
        }
    finally:
        db.close()


# =============================================================================
# CATALOG
# =============================================================================

@app.post("/api/catalog/update-from-page")
def update_catalog_from_page(page_path: str, current_user: User = Depends(get_current_user)):
    return ComponentCatalog().update_from_page(page_path)


@app.get("/api/catalog/list")
def list_catalog(current_user: User = Depends(get_current_user)):
    data = ComponentCatalog().get_all()
    return {
        "status": "success",
        "total_components": len(data.get("components", {})),
        "components": data.get("components", {}),
    }


class FieldSelection(BaseModel):
    resourceType: str
    version: str
    fields: List[str]
    label: Optional[str] = None


class TemplateFromCatalogRequest(BaseModel):
    selections: List[FieldSelection]


@app.post("/api/catalog/generate-template")
def generate_template_from_catalog(
    request: TemplateFromCatalogRequest,
    current_user: User = Depends(get_current_user),
):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst["A1"] = "AEM Content Updater – Generated Template"
    ws_inst["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws_inst["A3"] = "Important Rules"
    ws_inst["A3"].font = Font(bold=True, size=12)
    ws_inst["A4"] = "1. Do NOT change the header row (field names). They are locked."
    ws_inst["A5"] = "2. Only fill data rows. Leave a cell empty if you do not want to change that field."
    ws_inst["A6"] = "3. Page Path must start with /content/"
    ws_inst["A7"] = "4. Instance column: use 1, 2, 3... for multiple instances of the same component on a page."
    ws_inst["A8"] = "5. Upload this file in the tool and always review the Preview before applying."
    ws_inst.column_dimensions["A"].width = 90

    for sel in request.selections:
        sheet_name = (sel.label or sel.resourceType.split("/")[-1])[:31]
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Page Path", "Instance"] + sel.fields
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
            cell.protection = Protection(locked=True)
        for row in range(2, 8):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col, value="")
                cell.border = thin_border
                cell.protection = Protection(locked=False)
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 12
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        ws.protection.sheet = True
        ws.protection.password = "aem"

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=AEM_Template_From_Catalog.xlsx"},
    )


class ComponentSelection(BaseModel):
    resourceType: str
    label: Optional[str] = None


class TemplateRequest(BaseModel):
    components: List[ComponentSelection]


@app.post("/api/template/generate")
def generate_excel_template_legacy(
    request: TemplateRequest,
    current_user: User = Depends(get_current_user),
):
    service = ExcelTemplateService()
    excel_file = service.generate_template([comp.dict() for comp in request.components])
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=AEM_Content_Template.xlsx"},
    )


# =============================================================================
# DICTIONARY
# =============================================================================

@app.get("/api/dictionary")
def api_get_dictionary(current_user: User = Depends(get_current_user)):
    data = load_dictionary()
    return {"status": "success", "components": list_components(data), "raw": data}


@app.put("/api/dictionary/field")
def api_update_dictionary_field(payload: dict, current_user: User = Depends(get_current_user)):
    rt = payload.get("resourceType")
    fn = payload.get("field_name")
    labels = payload.get("ca_labels") or []
    if not rt or not fn:
        return {"status": "error", "message": "resourceType and field_name are required"}
    return update_field_aliases(rt, fn, labels)


@app.post("/api/dictionary/component")
def api_upsert_dictionary_component(payload: dict, current_user: User = Depends(get_current_user)):
    rt = payload.get("resourceType")
    if not rt:
        return {"status": "error", "message": "resourceType is required"}
    return upsert_component(rt, payload.get("label") or rt, payload.get("fields") or {})


@app.post("/api/dictionary/sync-from-fields")
def api_sync_dictionary(payload: dict, current_user: User = Depends(get_current_user)):
    return sync_from_catalog_fields(
        payload.get("resourceType") or "",
        payload.get("label") or "",
        payload.get("field_names") or [],
    )


# =============================================================================
# TEMPLATE HISTORY + GENERATE
# =============================================================================

@app.post("/api/excel/generate-template")
def api_generate_excel_template(payload: dict, current_user: User = Depends(get_current_user)):
    """
    Returns application/vnd...spreadsheetml.sheet binary.
    If generation fails, returns JSON error (not a fake xlsx).
    """
    try:
        selections = payload.get("selections") or []
        include_seo = payload.get("include_seo", True)
        include_assets = payload.get("include_assets", True)
        include_pages = payload.get("include_pages", True)
        include_add = payload.get("include_components_add", True)
        name = payload.get("name") or f"Template {datetime.utcnow().strftime('%Y%m%d_%H%M')}"

        known = []
        try:
            parent = payload.get("template_parent_path") or "/content/we-retail/us/en"
            listing = PageService().list_allowed_templates(parent, walk_ancestors=True)
            known = listing.get("templates") or []
        except Exception:
            known = []

        allowed_for_ca = []
        try:
            ac_path = payload.get("allowed_components_page_path") or payload.get("template_parent_path")
            if ac_path:
                ac = ComponentAddService().get_allowed_components(ac_path)
                allowed_for_ca = ac.get("allowed_for_ca") or []
        except Exception:
            allowed_for_ca = []

        try:
            save_template(name, selections, include_seo=include_seo, source="dictionary")
        except Exception:
            pass

        content = generate_template(
            selections,
            include_seo=include_seo,
            include_assets=include_assets,
            include_pages=include_pages,
            include_components_add=include_add,
            known_templates=known,
            allowed_components=allowed_for_ca,
            default_template_name=payload.get("default_template_name") or "Content Page",
        )
        if not content or not content.startswith(b"PK"):
            return {"status": "error", "message": "Template generation produced invalid file"}

        filename = f"AEM_Update_Template_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(content)),
            },
        )
    except Exception as e:
        return {"status": "error", "message": f"Template generation failed: {e}"}


@app.get("/api/templates/history")
def api_list_template_history(current_user: User = Depends(get_current_user)):
    return {"status": "success", "templates": list_templates()}


@app.delete("/api/templates/history/{template_id}")
def api_delete_template_history(template_id: str, current_user: User = Depends(get_current_user)):
    return delete_template(template_id)


@app.post("/api/templates/history/{template_id}/download")
def api_download_previous_template(template_id: str, current_user: User = Depends(get_current_user)):
    tpl = get_template(template_id)
    if not tpl:
        return {"status": "error", "message": "Template not found"}
    mark_used(template_id)
    content = generate_template(
        tpl.get("selections") or [],
        include_seo=bool(tpl.get("include_seo")),
        include_assets=True,
        include_pages=True,
    )
    filename = f"{(tpl.get('name') or 'AEM_Template').replace(' ', '_')}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =============================================================================
# EXCEL BULK — component/SEO update
# =============================================================================

@app.post("/api/excel/preview")
async def preview_excel_updates(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        return {"status": "error", "message": "Please upload a valid Excel file (.xlsx)"}
    return preview_excel(await file.read())


@app.post("/api/excel/apply")
async def apply_excel_updates(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Component UPDATE + SEO only (skips Assets / Pages / Add sheets)."""
    if not file.filename.endswith((".xlsx", ".xls")):
        return {"status": "error", "message": "Please upload a valid Excel file (.xlsx)"}
    return apply_excel(
        await file.read(),
        performed_by=current_user.full_name or current_user.username,
    )


@app.post("/api/excel/bulk/preview")
async def excel_bulk_preview(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Full workbook preview: Assets → Pages → Add → Update."""
    if not file.filename.endswith((".xlsx", ".xls")):
        return {"status": "error", "message": "Please upload a valid Excel file (.xlsx)"}
    sid = request.headers.get("X-Bulk-Session-Id") or ""
    return orchestrate_preview(
        await file.read(),
        username=current_user.username,
        session_id=sid or None,
    )



@app.post("/api/excel/bulk/apply")
async def excel_bulk_apply(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Full workbook apply in order: Assets → Pages → Add components → SEO/Update."""
    if not file.filename.endswith((".xlsx", ".xls")):
        return {"status": "error", "message": "Please upload a valid Excel file (.xlsx)"}
    return orchestrate_apply(
        await file.read(),
        performed_by=current_user.full_name or current_user.username,
    )


# =============================================================================
# EXCEL BULK — Assets
# =============================================================================



@app.post("/api/excel/bulk/validate")
async def excel_bulk_validate(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Post-apply verification: DAM, pages, components, field values vs Excel."""
    if not file.filename.endswith((".xlsx", ".xls")):
        return {"status": "error", "message": "Please upload a valid Excel file (.xlsx)"}
    return build_validation_report(await file.read())


@app.post("/api/excel/bulk/validate/export")
async def excel_bulk_validate_export(
    file: UploadFile = File(...),
    format: str = "xlsx",
    current_user: User = Depends(get_current_user),
):
    """
    Export validation report.
    Query/form: format=xlsx|json
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        return {"status": "error", "message": "Please upload a valid Excel file (.xlsx)"}
    content = await file.read()
    report = build_validation_report(content)
    fmt = (format or "xlsx").lower().strip()
    if fmt == "json":
        import json
        raw = json.dumps(report, indent=2, default=str).encode("utf-8")
        return StreamingResponse(
            io.BytesIO(raw),
            media_type="application/json",
            headers={"Content-Disposition": 'attachment; filename="AEM_Bulk_Validation_Report.json"'},
        )
    xlsx = validation_report_to_xlsx(report)
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="AEM_Bulk_Validation_Report.xlsx"'},
    )


@app.post("/api/excel/bulk/session/clear")
async def excel_bulk_session_clear(
    request: Request,
    current_user: User = Depends(get_current_user),
):
    sid = request.headers.get("X-Bulk-Session-Id") or ""
    return clear_session(sid or None)


@app.post("/api/excel/bulk/session/mark-applied")
async def excel_bulk_session_mark(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    sid = request.headers.get("X-Bulk-Session-Id") or ""
    content = await file.read()
    return mark_applied(sid or None, content, username=current_user.username)


@app.get("/api/excel/bulk/session")
async def excel_bulk_session_get(
    request: Request,
    current_user: User = Depends(get_current_user),
):
    sid = request.headers.get("X-Bulk-Session-Id") or ""
    s = get_session(sid or None)
    if not s:
        return {"status": "empty", "message": "No active bulk session (page reload starts fresh)"}
    return {"status": "success", "file_hash": s.get("file_hash"), "applied_at": s.get("applied_at")}

@app.post("/api/excel/components-add/preview")
async def excel_components_add_preview(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    from backend.app.services.excel_bulk_service import parse_add_sheets
    return parse_add_sheets(await file.read())


@app.post("/api/excel/components-add/apply")
async def excel_components_add_apply(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Apply only Add * sheets — skips pages that do not exist."""
    from collections import OrderedDict
    from backend.app.services.excel_bulk_service import parse_add_sheets
    from backend.app.services.component_add_service import ComponentAddService
    from backend.app.services.page_service import PageService

    content = await file.read()
    adds = parse_add_sheets(content)
    by_page = OrderedDict()
    for row in adds.get("rows") or []:
        by_page.setdefault(row["page_path"], []).append(row)
    add_svc = ComponentAddService()
    ps = PageService()
    add_results = []
    for page_path, rows in by_page.items():
        if not ps.path_exists(page_path):
            add_results.append({
                "page_path": page_path,
                "status": "skipped",
                "message": "Page does not exist — skipped component add",
                "components_requested": [r.get("component") for r in rows],
            })
            continue
        components = [
            {"component": r["component"], "properties": r.get("properties") or {}}
            for r in rows
        ]
        r = add_svc.apply_add(page_path, components)
        add_results.append({"page_path": page_path, "status": r.get("status"), "result": r})
    return {"status": "success", "pages": add_results, "parse_errors": adds.get("errors") or []}


@app.post("/api/excel/assets/preview")
async def excel_assets_preview(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    return plan_asset_uploads(await file.read())


@app.post("/api/excel/assets/apply")
async def excel_assets_apply(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    plan = plan_asset_uploads(content)
    if plan.get("status") != "success":
        return {**plan, "status": "error"}
    dam = DamService()
    results = []
    for p in plan.get("plans") or []:
        if p.get("errors"):
            results.append({
                "row": p.get("excel_row"),
                "status": "error",
                "source_path": p.get("source_path"),
                "target_path": p.get("target_path"),
                "errors": p["errors"],
            })
            continue
        up = dam.upload_from_local(
            page_dam_path=p["target_path"],
            local_page_folder=p["source_path"],
            confirm_create_folders=True,
        )
        st = up.get("status") or "error"
        results.append({
            "row": p.get("excel_row"),
            "status": st if st in ("success", "partial") else "error",
            "source_path": p.get("source_path"),
            "target_path": p.get("target_path"),
            "mode": up.get("mode") or p.get("mode"),
            "upload": up,
        })
    err_n = sum(1 for r in results if r.get("status") == "error")
    ok_n = sum(1 for r in results if r.get("status") == "success")
    overall = "success" if err_n == 0 else ("partial" if ok_n else "error")
    return {
        "status": overall,
        "message": f"Assets apply — ok: {ok_n}, errors: {err_n}",
        "results": results,
        "plan_summary": plan.get("summary"),
    }


# =============================================================================
# EXCEL BULK — Pages
# =============================================================================

@app.post("/api/excel/pages/preview")
async def excel_pages_preview(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    return preview_pages(await file.read())


@app.post("/api/excel/pages/apply")
async def excel_pages_apply(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    return apply_pages(
        await file.read(),
        performed_by=current_user.full_name or current_user.username,
    )


# =============================================================================
# DAM
# =============================================================================

@app.post("/api/dam/inspect")
def dam_inspect(
    page_dam_path: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user),
):
    return DamService().inspect_page_dam_path(page_dam_path)


@app.post("/api/dam/ensure-folders")
def dam_ensure_folders(
    page_dam_path: str = Body(...),
    confirm_create: bool = Body(False),
    current_user: User = Depends(get_current_user),
):
    return DamService().ensure_page_dam_structure(page_dam_path, confirm_create=confirm_create)


@app.post("/api/dam/scan-local")
def dam_scan_local(
    local_page_folder: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user),
):
    return DamService().scan_local_page_folder(local_page_folder)


@app.post("/api/dam/upload-local")
def dam_upload_local(
    page_dam_path: str = Body(...),
    local_page_folder: str = Body(...),
    confirm_create_folders: bool = Body(False),
    current_user: User = Depends(get_current_user),
):
    return DamService().upload_from_local(
        page_dam_path=page_dam_path,
        local_page_folder=local_page_folder,
        confirm_create_folders=confirm_create_folders,
    )


# =============================================================================
# PAGE CREATE (interactive UI)
# =============================================================================

@app.post("/api/page/inspect")
def page_inspect(
    target_path: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user),
):
    return PageService().inspect_page_path(target_path)


@app.post("/api/page/plan")
def page_plan(
    target_path: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user),
):
    return PageService().create_path_plan(target_path)


@app.post("/api/page/templates")
def page_templates(
    parent_path: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user),
):
    return PageService().list_allowed_templates(parent_path)


@app.post("/api/page/create")
def page_create(
    target_path: str = Body(...),
    steps: list = Body(...),
    default_title: str = Body(None),
    current_user: User = Depends(get_current_user),
):
    return PageService().execute_create(
        target_path=target_path,
        steps=steps,
        default_title=default_title,
    )


# =============================================================================
# COMPONENT ADD
# =============================================================================

@app.get("/api/page/containers")
def page_containers(page_path: str, current_user: User = Depends(get_current_user)):
    return ComponentAddService().discover_containers(page_path)


@app.get("/api/page/allowed-components")
def page_allowed_components(
    page_path: str,
    current_user: User = Depends(get_current_user),
):
    """CA-friendly allowed component names for this page container."""
    return ComponentAddService().get_allowed_components(page_path)


@app.post("/api/page/resolve-component")
def resolve_component_name(payload: dict, current_user: User = Depends(get_current_user)):
    """Body: { "component": "Title", "page_path": "/content/..." }"""
    return ComponentAddService().resolve_resource_type(
        payload.get("component") or payload.get("name") or "",
        payload.get("page_path"),
    )


@app.post("/api/page/components/add/preview")
def components_add_preview(payload: dict, current_user: User = Depends(get_current_user)):
    return ComponentAddService().plan_add(
        payload.get("page_path") or "",
        payload.get("components") or [],
    )


@app.post("/api/page/components/add")
def components_add_apply(payload: dict, current_user: User = Depends(get_current_user)):
    return ComponentAddService().apply_add(
        payload.get("page_path") or "",
        payload.get("components") or [],
    )


# =============================================================================
# MCP
# =============================================================================

mcp = FastApiMCP(
    app,
    name="AEM Content Updater MCP",
    description=(
        "Enterprise tools to discover AEM components, read dialog fields, "
        "safely update content, and view audit logs. All updates are audited."
    ),
)
mcp.mount_http()
